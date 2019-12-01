
# part db
from openpyxl import Workbook
import pickle
import traceback
# part html
import dash
from dash.dependencies import Input, Output
import dash_table
import dash_core_components as dcc
import dash_html_components as html
import pandas as pd



dataDir = "../data/"

def getCatTab(idlist, at, ca, pr):
    wb = Workbook()
    ws = wb.active
    ws.title = "Категории"
    for i in range(len(idlist)):
        ws.cell(row=i + 1, column=1).value = idlist[i]
        ws.cell(row=i + 1, column=2).value = ca.get(idlist[i], {'Name': 'Нет'})['Name']

    for id in idlist:
        try:
            id = str(id)
            atHead = dict((k, v['name']) for k, v in at.items() if v['categoryId'] == id)
            prList = dict((k, v) for k, v in pr.items() if v['CategoryID'] == id)
            caItem = ca.get(id, {'Name': 'Нет'})

            atListK = list(atHead.keys())
            atList = list(map(lambda x: atHead[x], atListK))
            atHead = list(atHead.items())
            gen = ['PositionID', 'CategoryID', 'ProductPositionName', 'ProductPositionOKPD2'] + atList
            caName = caItem['Name'].replace('/', " и ")

            listName = caName[:23] + " " + str(len(prList))
            ws = wb.create_sheet(listName)

            for i in range(len(gen)):
                ws.cell(row=1, column=i + 1).value = gen[i]
            # запись таблицы

            prod = list(prList.keys())
            for i in range(len(prod)):
                id = prod[i]
                item = prList[id]
                j = 1
                ws.cell(row=i + 2, column=j).value = id
                j += 1
                ws.cell(row=i + 2, column=j).value = caName
                j += 1
                ws.cell(row=i + 2, column=j).value = item['ProductPositionName']
                j += 1
                ws.cell(row=i + 2, column=j).value = item['ProductPositionOKPD2']
                j += 1
                for ak in range(len(atListK)):
                    val = item['Attributes'].get(atListK[ak])
                    ws.cell(row=i + 2, column=ak+j).value = "" if val == 'NULL' else val
        except:
            print(id)
            print('Ошибка:\n', traceback.format_exc())
    wb.save('alltables.xlsx')


def getCategories(product):
    cat = set()
    for i in product:
        cat.add(product[i]['CategoryID'])
    return list(cat)


if __name__== '__main__':
    with open(dataDir + 'data.bin', 'rb') as f:
        attributes = pickle.load(f)
        categories = pickle.load(f)
        product = pickle.load(f)
    allCat = getCategories(product)
    getCatTab(allCat, attributes, categories, product)



# часть с выводом в таблицы и графики



# df = pd.read_pickle.load(f)
    # read_csv('https://raw.githubusercontent.com/plotly/datasets/master/gapminder2007.csv')

with open(dataDir + 'data.bin', 'rb') as f:
    attributes = pickle.load(f)
    categories = pickle.load(f)
    product = pickle.load(f)
allCat = getCategories(product)
getCatTab(allCat, attributes, categories, product)

print(getCatTab())

app = dash.Dash(__name__)

app.layout = html.Div([
    dash_table.DataTable(
        id='datatable-interactivity',
        columns=[
            {"name": i, "id": i, "deletable": True, "selectable": True} for i in df.columns
        ],
        data=df.to_dict('records'),
        editable=True,
        filter_action="native",
        sort_action="native",
        sort_mode="multi",
        column_selectable="single",
        row_selectable="multi",
        row_deletable=True,
        selected_columns=[],
        selected_rows=[],
        page_action="native",
        page_current=0,
        page_size=10,
    ),
    html.Div(id='datatable-interactivity-container')
])


@app.callback(
    Output('datatable-interactivity', 'style_data_conditional'),
    [Input('datatable-interactivity', 'selected_columns')]
)
def update_styles(selected_columns):
    return [{
        'if': {'column_id': i},
        'background_color': '#D2F3FF'
    } for i in selected_columns]


@app.callback(
    Output('datatable-interactivity-container', "children"),
    [Input('datatable-interactivity', "derived_virtual_data"),
     Input('datatable-interactivity', "derived_virtual_selected_rows")])
def update_graphs(rows, derived_virtual_selected_rows):
    # When the table is first rendered, `derived_virtual_data` and
    # `derived_virtual_selected_rows` will be `None`. This is due to an
    # idiosyncracy in Dash (unsupplied properties are always None and Dash
    # calls the dependent callbacks when the component is first rendered).
    # So, if `rows` is `None`, then the component was just rendered
    # and its value will be the same as the component's dataframe.
    # Instead of setting `None` in here, you could also set
    # `derived_virtual_data=df.to_rows('dict')` when you initialize
    # the component.
    if derived_virtual_selected_rows is None:
        derived_virtual_selected_rows = []

    dff = df if rows is None else pd.DataFrame(rows)

    colors = ['#7FDBFF' if i in derived_virtual_selected_rows else '#0074D9'
              for i in range(len(dff))]

    return [
        dcc.Graph(
            id=column,
            figure={
                "data": [
                    {
                        "x": dff["id"],
                        "y": getCategories[id],
                        "type": Categories[id],
                        "marker": {"color": colors},
                    }
                ],
                "layout": {
                    "xaxis": {"automargin": True},
                    "yaxis": {
                        "automargin": True,
                        "title": {"text": column}
                    },
                    "height": 250,
                    "margin": {"t": 10, "l": 10, "r": 10},
                },
            },
        )
        # check if column exists - user may have deleted it
        # If `column.deletable=False`, then you don't
        # need to do this check.
        for column in ["pop2", "lifeExp4", "gdpPercap6"] if column in dff
    ]


if __name__ == '__main__':
    app.run_server(debug=True)

df = pd.read_csv('https://raw.githubusercontent.com/plotly/datasets/master/gapminder2007.csv')

app = dash.Dash(__name__)

app.layout = html.Div([
    dash_table.DataTable(
        id='datatable-interactivity',
        columns=[
            {"name": i, "id": i, "deletable": True, "selectable": True} for i in df.columns
        ],
        data=df.to_dict('records'),
        editable=True,
        filter_action="native",
        sort_action="native",
        sort_mode="multi",
        column_selectable="single",
        row_selectable="multi",
        row_deletable=True,
        selected_columns=[],
        selected_rows=[],
        page_action="native",
        page_current=0,
        page_size=10,
    ),
    html.Div(id='datatable-interactivity-container')
])


@app.callback(
    Output('datatable-interactivity', 'style_data_conditional'),
    [Input('datatable-interactivity', 'selected_columns')]
)
def update_styles(selected_columns):
    return [{
        'if': {'column_id': i},
        'background_color': '#D2F3FF'
    } for i in selected_columns]


@app.callback(
    Output('datatable-interactivity-container', "children"),
    [Input('datatable-interactivity', "derived_virtual_data"),
     Input('datatable-interactivity', "derived_virtual_selected_rows")])
def update_graphs(rows, derived_virtual_selected_rows):
    # When the table is first rendered, `derived_virtual_data` and
    # `derived_virtual_selected_rows` will be `None`. This is due to an
    # idiosyncracy in Dash (unsupplied properties are always None and Dash
    # calls the dependent callbacks when the component is first rendered).
    # So, if `rows` is `None`, then the component was just rendered
    # and its value will be the same as the component's dataframe.
    # Instead of setting `None` in here, you could also set
    # `derived_virtual_data=df.to_rows('dict')` when you initialize
    # the component.
    if derived_virtual_selected_rows is None:
        derived_virtual_selected_rows = []

    dff = df if rows is None else pd.DataFrame(rows)

    colors = ['#7FDBFF' if i in derived_virtual_selected_rows else '#0074D9'
              for i in range(len(dff))]

    return [
        dcc.Graph(
            id=column,
            figure={
                "data": [
                    {
                        "x": dff["id"],
                        "y": getCategories[id],
                        "type": "bar",
                        "marker": {"color": colors},
                    }
                ],
                "layout": {
                    "xaxis": {"automargin": True},
                    "yaxis": {
                        "automargin": True,
                        "title": {"text": column}
                    },
                    "height": 250,
                    "margin": {"t": 10, "l": 10, "r": 10},
                },
            },
        )
        # check if column exists - user may have deleted it
        # If `column.deletable=False`, then you don't
        # need to do this check.
        for column in ["pop", "lifeExp", "gdpPercap"] if column in dff
    ]


if __name__ == '__main__':
    app.run_server(debug=True)
