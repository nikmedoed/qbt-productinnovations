from openpyxl import Workbook
import pickle
import traceback

dataDir = "../data/"

def getCatTab(idlist, at, ca, pr):
    wb = Workbook()
    ws = wb.active
    ws.title = "Категории"
    for i in range(len(idlist)):
        ws.cell(row=i + 1, column=1).value = idlist[i]
        ws.cell(row=i + 1, column=2).value = ca.get(idlist[i], {'Name': 'Нет'})['Name']

    for id in idlist:
        try:
            id = str(id)
            atHead = dict((k, v['name']) for k, v in at.items() if v['categoryId'] == id)
            prList = dict((k, v) for k, v in pr.items() if v['CategoryID'] == id)
            caItem = ca.get(id, {'Name': 'Нет'})

            atListK = list(atHead.keys())
            atList = list(map(lambda x: atHead[x], atListK))
            atHead = list(atHead.items())
            gen = ['PositionID', 'CategoryID', 'ProductPositionName', 'ProductPositionOKPD2'] + atList
            caName = caItem['Name'].replace('/', " и ")

            listName = caName[:23] + " " + str(len(prList))
            ws = wb.create_sheet(listName)

            for i in range(len(gen)):
                ws.cell(row=1, column=i + 1).value = gen[i]
            # запись таблицы

            prod = list(prList.keys())
            for i in range(len(prod)):
                id = prod[i]
                item = prList[id]
                j = 1
                ws.cell(row=i + 2, column=j).value = id
                j += 1
                ws.cell(row=i + 2, column=j).value = caName
                j += 1
                ws.cell(row=i + 2, column=j).value = item['ProductPositionName']
                j += 1
                ws.cell(row=i + 2, column=j).value = item['ProductPositionOKPD2']
                j += 1
                for ak in range(len(atListK)):
                    val = item['Attributes'].get(atListK[ak])
                    ws.cell(row=i + 2, column=ak+j).value = "" if val == 'NULL' else val
        except:
            print(id)
            print('Ошибка:\n', traceback.format_exc())
    # wb.save('alltables.xlsx')


def getCategories(product):
    cat = set()
    for i in product:
        cat.add(product[i]['CategoryID'])
    return list(cat)


if __name__== '__main__':
    with open(dataDir + 'data.bin', 'rb') as f:
        attributes = pickle.load(f)
        categories = pickle.load(f)
        product = pickle.load(f)
    allCat = getCategories(product)
    getCatTab(allCat, attributes, categories, product)


# некрасиво но можно после инициализировать



import dash
import dash_html_components as html
import dash_table
import pandas as pd
from collections import OrderedDict


app = dash.Dash(__name__)

df = pd.DataFrame(OrderedDict([
    ('climate', ['Sunny', 'Snowy', 'Sunny', 'Rainy']),
    ('temperature', [13, 43, 50, 30]),
    ('city', ['NYC', 'Montreal', 'Miami', 'NYC'])
]))


app.layout = html.Div([
    dash_table.DataTable(
        id='table-dropdown',
        data=df.to_dict('records'),
        columns=[
            {'id': 'climate', 'name': 'climate', 'presentation': 'dropdown'},
            {'id': 'temperature', 'name': 'temperature'},
            {'id': 'city', 'name': 'city', 'presentation': 'dropdown'},
        ],

        editable=True,
        dropdown={
            'climate': {
                'options': [
                    {'label': i, 'value': i}
                    for i in df['climate'].unique()
                ]
            },
            'city': {
                 'options': [
                    {'label': i, 'value': i}
                    for i in df['city'].unique()
                ]
            }
        }
    ),
    html.Div(id='table-dropdown-container')
])


if __name__ == '__main__':
    app.run_server(debug=True)