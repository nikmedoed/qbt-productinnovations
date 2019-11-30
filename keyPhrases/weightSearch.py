from rake_nltk import Rake
import openpyxl


def getInn(path = "../data/inn.xlsx"):
    res = []
    wb = openpyxl.load_workbook(path).active
    names = []
    for i in range(wb.max_column):
        v = wb.cell(row=1, column=i+1).value
        if v: names.append(v)
    for i in range(wb.max_row-2):
        r = i+2
        temp = {}
        for i in range(len(names)):
            v = wb.cell(row=r, column=2 + i).value
            if v: temp[names[i]] = v
        res.append(temp)
    return res

def analyse(tab):
    keys = ['Наименование продукта/технологии', 'Уникальные характеристики', 'Задачи, которые решает продукт', 'Технические характеристики', 'Ожидаемые эффекты']
    res = []
    r = Rake()
    r.language = "russian"
    # Extraction given the text.
    for i in tab:
        text = "\n".join(list(map(lambda x: i[x], keys)))
        r.extract_keywords_from_text(text)
        ranked = r.get_ranked_phrases_with_scores()
        res.append(ranked)
    return res

if __name__ == '__main__':
    path = "../data/inn.xlsx"
    table = getInn(path)
    weights = analyse(table)
    c = 0
    s = 0
    k = 2
    book = openpyxl.load_workbook(path)
    wb = book.active
    col = 13 # wb.max_column+1
    for i in weights:
        for j in i:
            c += 1
            s += j[0]
        temp = list(filter(lambda x: x[0] > 65, i))
        wb.cell(row=k, column=col).value = "\n".join([a[1] for a in temp])
        wb.cell(row=k, column=col+1).value = "\n".join([str(round(a[0], 2)) for a in temp])
        k+=1
    book.save(path)
    print (s/c)

from gensim.models import word2vec