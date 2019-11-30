from openpyxl import Workbook
import pickle
import traceback

dataDir = "data/"

def getCatTab(id, at, ca, pr):
    id = str(id)
    try:
        atHead = dict((k, v['name']) for k, v in at.items() if v['categoryId'] == id)
        prList = dict((k, v) for k, v in pr.items() if v['CategoryID'] == id)
        caItem = ca.get(id, {'Name': 'Нет'})

        atListK = list(atHead.keys())
        atList = list(map(lambda x: atHead[x], atListK))
        atHead = list(atHead.items())
        gen = ['PositionID', 'CategoryID', 'ProductPositionName', 'ProductPositionOKPD2'] + atList
        caName = caItem['Name'].replace('/', " и ")

        wb = Workbook()
        ws = wb.active
        ws.title = caName[:31]

        for i in range(len(gen)):
            ws.cell(row=1, column=i + 1).value = gen[i]
        # запись таблицы

        prod = list(prList.keys())
        for i in range(len(prod)):
            id = prod[i]
            item = prList[id]
            j = 1
            ws.cell(row=i + 2, column=j).value = id
            j += 1
            ws.cell(row=i + 2, column=j).value = caName
            j += 1
            ws.cell(row=i + 2, column=j).value = item['ProductPositionName']
            j += 1
            ws.cell(row=i + 2, column=j).value = item['ProductPositionOKPD2']
            j += 1
            for ak in range(len(atListK)):
                val = item['Attributes'].get(atListK[ak])
                ws.cell(row=i + 2, column=ak+j).value = "" if val == 'NULL' else val
        wb.save('tables/' + str(len(prList)) + " i" + id + ' - ' + caName[:50] + '.xlsx')
    except:
        print(id)
        print('Ошибка:\n', traceback.format_exc())

def getCategories(product):
    cat = set()
    for i in product:
        cat.add(product[i]['CategoryID'])
    return list(cat)


if __name__== '__main__':
    with open(dataDir + 'data.bin', 'rb') as f:
        attributes = pickle.load(f)
        categories = pickle.load(f)
        product = pickle.load(f)
    allCat = getCategories(product)
    for i in allCat:
        getCatTab(i, attributes, categories, product)


