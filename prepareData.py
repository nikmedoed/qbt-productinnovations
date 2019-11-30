import openpyxl
import csv
import pickle


dataDir = "data/"
dataAttributes = dataDir + "att.xlsx"
dataCategories = dataDir + "cat.xlsx"
dataProduct = dataDir + "prod.csv"

def getAttr(path = dataAttributes):
    res = dict()
    wb = openpyxl.load_workbook(path).active
    for i in range(wb.max_row-2):
        r = i+2
        res[wb.cell(row=r, column=6).value] = {
            'name': wb.cell(row=r, column=7).value,
            'type': wb.cell(row=r, column=8).value,
            'categoryId': wb.cell(row=r, column=1).value
        }
    return res

def getCat(path = dataCategories):
    res = dict()
    wb = openpyxl.load_workbook(path).active
    names = []
    for i in range(wb.max_column-1):
        v = wb.cell(row=1, column=i+1).value
        if v: names.append(v)
    for i in range(wb.max_row-2):
        r = i+2
        id = wb.cell(row=r, column=1).value
        res[id] = {}
        for i in range(len(names)):
            v = wb.cell(row=r, column=2 + i).value
            if v: res[id][names[i]] = v
    return res

def getProd(path = dataProduct):
    res = dict()
    reader = csv.DictReader(open(path, "r"), delimiter='\t')
    # head = reader.fieldnames
    for line in reader:
        d = dict(line)
        id = d.pop('PositionID')
        i = d.pop('ProductPositionAttributeID')
        v = d.pop('ProductPositionAttribute')
        if not (id in res):
            res[id] = d
            res[id]['Attributes'] = {}
        res[id]['Attributes'][i] = v
    return res

attributes = getAttr()
categories = getCat()
product = getProd()

with open(dataDir + 'data.bin', 'wb') as f:
    pickle.dump(attributes, f)
    pickle.dump(categories, f)
    pickle.dump(product, f)